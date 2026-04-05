"""
Q2: NLP Analysis on Barclays Tweets (BarclaysTweets_topics sheet)
Using Python (TextBlob, sklearn, spaCy-like NLP) instead of MeaningCloud Add-in.
Performs:
  1. Text Classification
  2. Topics Extraction
  3. Text Clustering
  4. Deep Categorization (Basel II risk categories)
"""

import pandas as pd
import numpy as np
from textblob import TextBlob
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
import re
from collections import Counter
import openpyxl

# Load data
df = pd.read_excel("Barclays_Tweets.xlsx", sheet_name="BarclaysTweets_topics")
tweets = df["tweet_text"].dropna().astype(str).tolist()
print(f"Total tweets loaded: {len(tweets)}")

# ============================================================
# 1. TEXT CLASSIFICATION (Sentiment-based + Subjectivity)
# ============================================================
print("\n" + "=" * 80)
print("1. TEXT CLASSIFICATION")
print("=" * 80)

classifications = []
for tweet in tweets:
    blob = TextBlob(tweet)
    pol = blob.sentiment.polarity
    subj = blob.sentiment.subjectivity

    # Classify by sentiment
    if pol > 0.1:
        sentiment_class = "Positive"
    elif pol < -0.1:
        sentiment_class = "Negative"
    else:
        sentiment_class = "Neutral"

    # Classify by subjectivity
    if subj > 0.5:
        subj_class = "Opinion/Subjective"
    else:
        subj_class = "Factual/Objective"

    classifications.append((sentiment_class, subj_class, pol, subj))

df_class = pd.DataFrame(
    classifications,
    columns=["Sentiment_Class", "Subjectivity_Class", "Polarity", "Subjectivity"],
)

print(f"\nSentiment Classification Distribution:")
for cls, count in df_class["Sentiment_Class"].value_counts().items():
    print(f"  {cls:<12}: {count:>4} tweets ({count/len(df_class)*100:.1f}%)")

print(f"\nSubjectivity Classification Distribution:")
for cls, count in df_class["Subjectivity_Class"].value_counts().items():
    print(f"  {cls:<20}: {count:>4} tweets ({count/len(df_class)*100:.1f}%)")

print(f"\nSample Classifications (first 10):")
print(f"{'#':<4} {'Sentiment':<12} {'Type':<22} {'Tweet (first 55 chars)'}")
print("-" * 93)
for i in range(min(10, len(tweets))):
    print(
        f"{i+1:<4} {df_class['Sentiment_Class'].iloc[i]:<12} "
        f"{df_class['Subjectivity_Class'].iloc[i]:<22} {tweets[i][:55]}"
    )

# ============================================================
# 2. TOPICS EXTRACTION (using TF-IDF for key terms)
# ============================================================
print("\n" + "=" * 80)
print("2. TOPICS EXTRACTION")
print("=" * 80)

# Extract key topics using TF-IDF
vectorizer = TfidfVectorizer(
    max_features=500,
    stop_words="english",
    ngram_range=(1, 2),
    min_df=2,
    max_df=0.95,
)
tfidf_matrix = vectorizer.fit_transform(tweets)
feature_names = vectorizer.get_feature_names_out()

# Get top terms across all documents
mean_tfidf = tfidf_matrix.mean(axis=0).A1
top_indices = mean_tfidf.argsort()[::-1][:30]

print(f"\nTop 30 Topics/Key Terms (by TF-IDF importance):")
print(f"{'Rank':<6} {'Topic/Term':<30} {'TF-IDF Score'}")
print("-" * 50)
for rank, idx in enumerate(top_indices, 1):
    print(f"{rank:<6} {feature_names[idx]:<30} {mean_tfidf[idx]:.4f}")

# Extract topics per tweet using noun phrases (TextBlob)
print(f"\nSample Topic Extraction per Tweet (first 10):")
print("-" * 80)
extracted_topics = []
for i, tweet in enumerate(tweets):
    blob = TextBlob(tweet)
    nouns = blob.noun_phrases
    topics = list(set(nouns))[:5] if nouns else ["N/A"]
    extracted_topics.append(", ".join(topics))
    if i < 10:
        print(f"Tweet {i+1}: {tweet[:60]}...")
        print(f"  Topics: {', '.join(topics)}")

# ============================================================
# 3. TEXT CLUSTERING (K-Means on TF-IDF)
# ============================================================
print("\n" + "=" * 80)
print("3. TEXT CLUSTERING")
print("=" * 80)

n_clusters = 5
kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
clusters = kmeans.fit_predict(tfidf_matrix)

print(f"\nNumber of clusters: {n_clusters}")
print(f"\nCluster Distribution:")
for c in range(n_clusters):
    count = (clusters == c).sum()
    print(f"  Cluster {c+1}: {count:>4} tweets ({count/len(tweets)*100:.1f}%)")

# Show top terms per cluster
print(f"\nTop 10 Terms per Cluster:")
order_centroids = kmeans.cluster_centers_.argsort()[:, ::-1]
for c in range(n_clusters):
    top_terms = [feature_names[idx] for idx in order_centroids[c, :10]]
    print(f"\n  Cluster {c+1}: {', '.join(top_terms)}")
    # Show sample tweets
    cluster_indices = np.where(clusters == c)[0][:3]
    for idx in cluster_indices:
        print(f"    - {tweets[idx][:70]}...")

# ============================================================
# 4. DEEP CATEGORIZATION (Basel II Risk Categories)
# ============================================================
print("\n" + "=" * 80)
print("4. DEEP CATEGORIZATION (Basel II Operational Risk Categories)")
print("=" * 80)

# Define keywords for Basel II event type categories
basel_categories = {
    "Internal Fraud": [
        "fraud", "misappropriation", "tax evasion", "bribery", "corruption",
        "embezzlement", "insider", "manipulation", "misconduct", "rogue",
        "unauthorized", "falsif", "forge"
    ],
    "External Fraud": [
        "hack", "theft", "phishing", "cyber", "breach", "stolen",
        "identity theft", "scam", "counterfeit", "forgery", "external fraud",
        "data breach", "ransomware", "malware"
    ],
    "Employment Practices": [
        "discrimination", "harassment", "worker", "employee", "workplace",
        "safety", "compensation", "labor", "termination", "diversity",
        "staff", "hiring", "fired"
    ],
    "Clients Products Business": [
        "customer", "client", "product", "service", "complaint", "mis-selling",
        "misselling", "fiduciary", "suitability", "advisory", "churning",
        "complaint", "consumer", "account", "fee", "charge", "rate",
        "interest", "loan", "mortgage", "credit"
    ],
    "Damage to Physical Assets": [
        "disaster", "flood", "fire", "earthquake", "terrorism", "vandalism",
        "damage", "physical", "hurricane", "storm"
    ],
    "Business Disruption": [
        "outage", "system failure", "disruption", "downtime", "crash",
        "glitch", "error", "bug", "software", "hardware", "IT failure",
        "technical", "server", "website", "app", "online", "mobile"
    ],
    "Execution Delivery Process": [
        "error", "mistake", "processing", "settlement", "delivery",
        "compliance", "regulatory", "reporting", "data entry", "accounting",
        "fine", "penalty", "regulation", "libor", "rigging"
    ],
}


def categorize_tweet(text):
    """Categorize tweet into Basel II risk categories."""
    text_lower = text.lower()
    scores = {}
    for category, keywords in basel_categories.items():
        score = sum(1 for kw in keywords if kw in text_lower)
        if score > 0:
            scores[category] = score
    if scores:
        return max(scores, key=scores.get)
    return "Uncategorized"


categories = [categorize_tweet(t) for t in tweets]

print(f"\nBasel II Category Distribution:")
cat_counts = Counter(categories)
for cat, count in cat_counts.most_common():
    print(f"  {cat:<35}: {count:>4} tweets ({count/len(tweets)*100:.1f}%)")

print(f"\nSample Categorizations (first 15):")
print(f"{'#':<4} {'Category':<35} {'Tweet (first 50 chars)'}")
print("-" * 90)
for i in range(min(15, len(tweets))):
    print(f"{i+1:<4} {categories[i]:<35} {tweets[i][:50]}")

# ============================================================
# Save results back to Excel
# ============================================================
wb = openpyxl.load_workbook("Barclays_Tweets.xlsx")
ws = wb["BarclaysTweets_topics"]

ws.cell(row=1, column=5).value = "Topic_Extracted"
ws.cell(row=1, column=6).value = "Text_Classification"
ws.cell(row=1, column=7).value = "Cluster"
ws.cell(row=1, column=8).value = "Basel_Category"

for i in range(len(tweets)):
    ws.cell(row=i + 2, column=5).value = extracted_topics[i] if i < len(extracted_topics) else ""
    ws.cell(row=i + 2, column=6).value = df_class["Sentiment_Class"].iloc[i] if i < len(df_class) else ""
    ws.cell(row=i + 2, column=7).value = int(clusters[i] + 1) if i < len(clusters) else ""
    ws.cell(row=i + 2, column=8).value = categories[i] if i < len(categories) else ""

wb.save("Barclays_Tweets.xlsx")
print(f"\nResults saved to 'Barclays_Tweets.xlsx' sheet 'BarclaysTweets_topics'")
