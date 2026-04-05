"""
Q1: Sentiment Analysis on Barclays Tweets
Using Python (TextBlob) instead of Azure ML Add-in.
Calculates positive, negative sentiments, and sentiment scores
using Relative Proportional Difference.
"""

import pandas as pd
from textblob import TextBlob
import openpyxl

# Load data from BarclaysTweets sheet
df = pd.read_excel("Barclays_Tweets.xlsx", sheet_name="BarclaysTweets")

print(f"Total tweets: {len(df)}")
print(f"Columns: {list(df.columns)}")
print(f"\nSample tweets:")
for i in range(min(3, len(df))):
    print(f"  {i+1}. {df['tweet_text'].iloc[i][:80]}...")


def analyze_sentiment(text):
    """Analyze sentiment using TextBlob and return positive, negative counts and RPD score."""
    if pd.isna(text) or str(text).strip() == "":
        return 0, 0, 0

    blob = TextBlob(str(text))
    # TextBlob polarity ranges from -1 to 1
    # We convert to positive/negative counts based on sentence-level analysis
    positive_count = 0
    negative_count = 0

    if len(blob.sentences) == 0:
        return 0, 0, 0

    for sentence in blob.sentences:
        if sentence.sentiment.polarity > 0:
            positive_count += 1
        elif sentence.sentiment.polarity < 0:
            negative_count += 1

    # If single sentence, use polarity directly
    if len(blob.sentences) == 1:
        pol = blob.sentiment.polarity
        if pol > 0:
            positive_count = round(abs(pol) * 10)
            negative_count = 0
        elif pol < 0:
            positive_count = 0
            negative_count = round(abs(pol) * 10)
        else:
            positive_count = 0
            negative_count = 0

    # Relative Proportional Difference: (R - L) / (R + L)
    # R = positive, L = negative
    total = positive_count + negative_count
    if total == 0:
        rpd_score = 0
    else:
        rpd_score = (positive_count - negative_count) / (positive_count + negative_count)

    return positive_count, negative_count, round(rpd_score, 4)


# Apply sentiment analysis
results = df["tweet_text"].apply(analyze_sentiment)
df["positive"] = results.apply(lambda x: x[0])
df["negative"] = results.apply(lambda x: x[1])
df["Sentiment_score"] = results.apply(lambda x: x[2])

# Display results
print("\n" + "=" * 80)
print("SENTIMENT ANALYSIS RESULTS")
print("=" * 80)

print(f"\n--- Summary Statistics ---")
print(f"Total tweets analyzed: {len(df)}")
print(f"Positive tweets (score > 0): {(df['Sentiment_score'] > 0).sum()}")
print(f"Negative tweets (score < 0): {(df['Sentiment_score'] < 0).sum()}")
print(f"Neutral tweets (score = 0):  {(df['Sentiment_score'] == 0).sum()}")
print(f"\nAverage Sentiment Score: {df['Sentiment_score'].mean():.4f}")
print(f"Std Dev of Sentiment Score: {df['Sentiment_score'].std():.4f}")

print(f"\n--- Sample Results (first 15 tweets) ---")
print(f"{'Row':<5} {'Positive':<10} {'Negative':<10} {'Score':<10} {'Tweet (first 60 chars)'}")
print("-" * 95)
for i in range(min(15, len(df))):
    tweet = str(df["tweet_text"].iloc[i])[:60]
    print(
        f"{i+1:<5} {df['positive'].iloc[i]:<10} {df['negative'].iloc[i]:<10} "
        f"{df['Sentiment_score'].iloc[i]:<10.4f} {tweet}"
    )

# Write results back to Excel
wb = openpyxl.load_workbook("Barclays_Tweets.xlsx")
ws = wb["BarclaysTweets"]

# Write headers if not present
ws.cell(row=1, column=5).value = "positive"
ws.cell(row=1, column=6).value = "negative"
ws.cell(row=1, column=7).value = "Sentiment_score"

for i in range(len(df)):
    ws.cell(row=i + 2, column=5).value = df["positive"].iloc[i]
    ws.cell(row=i + 2, column=6).value = df["negative"].iloc[i]
    ws.cell(row=i + 2, column=7).value = df["Sentiment_score"].iloc[i]

wb.save("Barclays_Tweets.xlsx")
print(f"\nResults saved to 'Barclays_Tweets.xlsx' sheet 'BarclaysTweets' columns E, F, G")

# Distribution of sentiments
print(f"\n--- Sentiment Distribution ---")
bins = [(-1.01, -0.5), (-0.5, -0.01), (0, 0), (0.01, 0.5), (0.5, 1.01)]
labels = ["Strong Negative", "Weak Negative", "Neutral", "Weak Positive", "Strong Positive"]
for (lo, hi), label in zip(bins, labels):
    if lo == 0 and hi == 0:
        count = (df["Sentiment_score"] == 0).sum()
    else:
        count = ((df["Sentiment_score"] >= lo) & (df["Sentiment_score"] <= hi)).sum()
    print(f"  {label:<20}: {count:>4} tweets ({count/len(df)*100:.1f}%)")
