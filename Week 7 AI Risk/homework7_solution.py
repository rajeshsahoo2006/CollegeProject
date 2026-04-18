"""
Homework 7 - AI in Technology Risk Assessment
BADM 566 Application of AI in Risk Management

Q1a: Risk Assessment - Extract topics from emails, calculate risk scores
Q1b: Linear Regression on normalized risk data
Q2:  SVM classification to predict risk factor relevance

Note: The Vaimal Excel Add-in and MeaningCloud Add-on could not be installed
due to compatibility issues with the current system environment. As an
alternative, Python (with scikit-learn, pandas, and openpyxl) was used to
perform the same analyses:
  - Topic extraction from unstructured email data (replacing MeaningCloud)
    using keyword-based NLP matching against the Risk Matrix taxonomy
  - Linear Regression (replacing Vaimal's regression tool) using
    scikit-learn's LinearRegression with Min-Max normalized features
  - SVM Classification (replacing Vaimal's SVM tool) using scikit-learn's
    SVC with RBF kernel
All results are exported to 'Week7-Risk_Assessment_Results.xlsx' with the
same sheet structure (Risk Assessment, Normalized, Linear Regression,
TechTrain, TechTest, Predict) as required by the homework.
"""

import numpy as np
import pandas as pd
import re
from sklearn.linear_model import LinearRegression
from sklearn.svm import SVC
from sklearn.preprocessing import LabelEncoder, MinMaxScaler
from sklearn.metrics import classification_report, accuracy_score, confusion_matrix
import openpyxl
import warnings
warnings.filterwarnings('ignore')

# =============================================================================
# Q1a: Risk Assessment - Topic Extraction & Risk Score Calculation
# =============================================================================
print("=" * 70)
print("Q1a: Technology Risk Assessment - Topic Extraction & Risk Scoring")
print("=" * 70)

# Load the emails from the Excel file
wb = openpyxl.load_workbook(
    'Week7-Risk Assessment.xlsx', data_only=True
)

# Read emails
ws_data = wb['week7-data']
emails = []
for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, values_only=True):
    if row[0]:
        emails.append(str(row[0]).strip())

print(f"\nTotal emails to analyze: {len(emails)}")
print("-" * 70)

# Define the risk taxonomy based on the Risk Matrix sheet
risk_taxonomy = {
    'Software failure': {
        'Testing': {'impact': 4, 'severity': 'High', 'risk_control': 'Yes',
                    'keywords': ['test', 'testing', 'tested', 'test problems']},
        'Design': {'impact': 3, 'severity': 'Medium', 'risk_control': 'Yes',
                   'keywords': ['design', 'bad design']},
        'End of Life (EOL)': {'impact': 3, 'severity': 'Medium', 'risk_control': 'No',
                              'keywords': ['end-of-life', 'eol', 'end of life']},
        'End of Vendor Support (EOVS)': {'impact': 2, 'severity': 'Low', 'risk_control': 'Yes',
                                         'keywords': ['vendor support', 'eovs', 'end of vendor support']},
        'Vulnerbility': {'impact': 3, 'severity': 'Medium', 'risk_control': 'Yes',
                         'keywords': ['vulnerability', 'vulnerbility', 'buggy', 'bug']},
        'Hacking/Malware': {'impact': 4, 'severity': 'High', 'risk_control': 'Yes',
                            'keywords': ['hack', 'malware', 'hacking', 'virus']},
    },
    'Hardware Failure': {
        'Natural/Man-made disasters': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'No',
                                       'keywords': ['fire', 'flood', 'disaster', 'short-circuit']},
        'Hard drive failures': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'No',
                                'keywords': ['disk', 'hard drive', 'disk drive']},
        'Network failures': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'Yes',
                             'keywords': ['network failure', 'network']},
        'Loss of Power due to Insufficient Backup': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'No',
                                                      'keywords': ['power', 'backup power', 'insufficient backup']},
    },
    'Users & computers': {
        'User accounts with passwords never expire.': {'impact': 3, 'severity': 'Medium', 'risk_control': 'No',
                                                        'keywords': ['never-expire', 'never expire', 'passwords category never']},
        'User accounts with passwords not required to expire.': {'impact': 2, 'severity': 'Low', 'risk_control': 'Yes',
                                                                  'keywords': ['expire-not-required', 'not required to expire', 'expire not required']},
        'Inactive User accounts': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'No',
                                   'keywords': ['inactive user', 'inactive']},
    },
    'Permissions': {
        'User accounts with administrative permissions': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'No',
                                                           'keywords': ['administrative permissions', 'admin permissions']},
        'Empty security groups.': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'No',
                                   'keywords': ['empty security groups', 'security groups']},
        'Administrative Groups': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'No',
                                  'keywords': ['administrative groups']},
    },
    'Data': {
        'Shared folders accessible by everyone': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'No',
                                                   'keywords': ['shared folders', 'accessible by everyone']},
        'File names containing sensitive data.': {'impact': 5, 'severity': 'Catastrophic', 'risk_control': 'No',
                                                   'keywords': ['sensitive data', 'file names containing']},
    },
}

# Probability mapping from text
probability_map = {
    'unlikely': 1, 'rare': 2, 'possible': 3, 'likely': 4, 'certain': 5
}

# Extract probability from email text (percentage)
def extract_probability(text):
    """Extract probability from text. Default to 'Likely' (4) if not found."""
    match = re.search(r'(\d+)%', text)
    if match:
        pct = int(match.group(1))
        if pct <= 15:
            return 1  # Unlikely
        elif pct <= 40:
            return 2  # Rare
        elif pct <= 60:
            return 3  # Possible
        elif pct <= 80:
            return 4  # Likely
        else:
            return 5  # Certain
    return 4  # Default to "Likely" as per instructions


# Extract topics from emails - allow multiple topics per email
extracted_events = []

for i, email in enumerate(emails):
    email_lower = email.lower()
    probability = extract_probability(email)
    found_any = False
    for event_type, sub_events in risk_taxonomy.items():
        for sub_event, info in sub_events.items():
            for keyword in info['keywords']:
                if keyword.lower() in email_lower:
                    extracted_events.append({
                        'Email_Index': i + 1,
                        'Email': email[:80] + '...' if len(email) > 80 else email,
                        'Event': event_type,
                        'Sub-Event': sub_event,
                        'Probability': probability,
                        'Risk Control Present': info['risk_control'],
                        'Impact_Category': info['impact'],
                        'Severity': info['severity'],
                    })
                    found_any = True
                    break  # only break keyword loop, continue to next sub-event

print("\nExtracted Events from Emails:")
print("-" * 70)
df_extracted = pd.DataFrame(extracted_events)
for _, row in df_extracted.iterrows():
    print(f"  Email {row['Email_Index']}: Event={row['Event']}, "
          f"Sub-Event={row['Sub-Event']}, Prob={row['Probability']}")

# Count events for each category
print("\n\nEvent Counts by Category:")
print("-" * 70)
event_counts = df_extracted.groupby('Event').size()
sub_event_counts = df_extracted.groupby(['Event', 'Sub-Event']).size().reset_index(name='Count')

for event_type in event_counts.index:
    total = event_counts[event_type]
    print(f"\n  {event_type}: Total = {total}")
    subs = sub_event_counts[sub_event_counts['Event'] == event_type]
    for _, sub in subs.iterrows():
        print(f"    - {sub['Sub-Event']}: {sub['Count']}")

# Build the risk assessment table
print("\n\nRisk Assessment Table:")
print("-" * 70)
print(f"{'Event':<20} {'Sub-Event':<45} {'Count':>5} {'Prob':>5} "
      f"{'Control':>7} {'L':>3} {'I':>6} {'Risk':>6}")
print("-" * 70)

risk_rows = []
for _, sub in sub_event_counts.iterrows():
    event = sub['Event']
    sub_event = sub['Sub-Event']
    count = sub['Count']
    total_event = event_counts[event]

    # Get info from taxonomy
    info = risk_taxonomy[event][sub_event]

    # Get probability (use the extracted one or default)
    prob_rows = df_extracted[(df_extracted['Event'] == event) &
                             (df_extracted['Sub-Event'] == sub_event)]
    probability = prob_rows['Probability'].values[0]
    risk_control = info['risk_control']

    # Calculate Likelihood based on probability
    # From the example: likelihood ranges 1-5
    likelihood = probability

    # Calculate Impact = Impact_Category * (sub_event_count / total_event_count)
    impact = info['impact'] * (count / total_event)

    # Risk factor = Likelihood x Impact
    risk_factor = likelihood * impact

    risk_rows.append({
        'Event': event,
        'Sub-Event': sub_event,
        "Sub-Events' Counts": count,
        'Probability': probability,
        'Risk Control Present': risk_control,
        'Likelihood (L)': likelihood,
        'Impact (I)': round(impact, 2),
        'Risk factor (L x I)': round(risk_factor, 2),
    })

    print(f"{event:<20} {sub_event:<45} {count:>5} {probability:>5} "
          f"{risk_control:>7} {likelihood:>3} {impact:>6.2f} {risk_factor:>6.2f}")

df_risk = pd.DataFrame(risk_rows)

# Now build the full 18-row table matching the Normalized sheet structure
# Use the template from the Normalized sheet for all sub-events
print("\n\nFull Risk Matrix (matching Normalized sheet format):")
print("-" * 70)

# Build complete risk table using ALL sub-events from taxonomy
full_risk_rows = []
for event_type, sub_events in risk_taxonomy.items():
    # Get total count for this event type from extracted data
    total_event = event_counts.get(event_type, 0)

    for sub_event_name, info in sub_events.items():
        # Get count for this sub-event
        match = sub_event_counts[
            (sub_event_counts['Event'] == event_type) &
            (sub_event_counts['Sub-Event'] == sub_event_name)
        ]
        count = int(match['Count'].values[0]) if len(match) > 0 else 1

        # Get probability from extracted emails, or default to 4 (Likely)
        prob_rows = df_extracted[
            (df_extracted['Event'] == event_type) &
            (df_extracted['Sub-Event'] == sub_event_name)
        ]
        probability = int(prob_rows['Probability'].values[0]) if len(prob_rows) > 0 else 4

        risk_control = info['risk_control']

        # Likelihood = probability rating
        likelihood = probability

        # Impact = Impact_Category * (count / total_event_count)
        if total_event > 0:
            impact = info['impact'] * (count / total_event)
        else:
            impact = info['impact'] * (count / max(count, 1))

        risk_factor = likelihood * impact

        full_risk_rows.append({
            'Event': event_type,
            'Sub-Event': sub_event_name,
            "Sub-Events' Counts": count,
            'Probability': probability,
            'Risk Control Present': risk_control,
            'Likelihood (L)': likelihood,
            'Impact (I)': round(impact, 2),
            'Risk factor (L x I)': round(risk_factor, 2),
        })

df_full_risk = pd.DataFrame(full_risk_rows)
print(df_full_risk.to_string(index=False))

# =============================================================================
# Q1b: Linear Regression on Normalized Data
# =============================================================================
print("\n\n" + "=" * 70)
print("Q1b: Linear Regression Analysis")
print("=" * 70)

# Prepare data following the Linear Regression example sheet format
# Convert categorical variables to numeric IDs
event_le = LabelEncoder()
sub_event_le = LabelEncoder()

df_reg = df_full_risk.copy()

# Encode Event -> Event ID (1-based)
unique_events = df_reg['Event'].unique()
event_id_map = {e: i+1 for i, e in enumerate(unique_events)}
df_reg['Event ID'] = df_reg['Event'].map(event_id_map)

# Encode Sub-Event -> Sub-Event ID (1-based)
df_reg['Sub-Event ID'] = range(1, len(df_reg) + 1)

# Encode Risk Control Present: Yes=1, No=0
df_reg['Risk Control Present Num'] = df_reg['Risk Control Present'].map({'Yes': 1, 'No': 0})

print("\nEvent ID Mapping:")
for event, eid in event_id_map.items():
    print(f"  {eid}: {event}")

print("\nPrepared Data for Regression:")
reg_cols = ['Event ID', 'Sub-Event ID', "Sub-Events' Counts", 'Probability',
            'Risk Control Present Num', 'Likelihood (L)', 'Impact (I)', 'Risk factor (L x I)']
df_reg_display = df_reg[reg_cols].copy()
df_reg_display.columns = ['Event ID', 'Sub-Event ID', "Sub-Events' Counts", 'Probability',
                           'Risk Control', 'Likelihood (L)', 'Impact (I)', 'Risk factor (L x I)']
print(df_reg_display.to_string(index=False))

# Normalize the features using Min-Max scaling
features = ['Event ID', 'Sub-Event ID', "Sub-Events' Counts", 'Probability',
            'Risk Control Present Num', 'Likelihood (L)', 'Impact (I)']
target = 'Risk factor (L x I)'

X = df_reg[features].values.astype(float)
y = df_reg[target].values.astype(float)

# Normalize features
scaler = MinMaxScaler()
X_normalized = scaler.fit_transform(X)

print("\nNormalized Features:")
df_norm = pd.DataFrame(X_normalized, columns=features)
df_norm['Risk factor'] = y
print(df_norm.round(4).to_string(index=False))

# Fit Linear Regression
reg = LinearRegression()
reg.fit(X_normalized, y)

print("\n--- Linear Regression Results ---")
print(f"Intercept: {reg.intercept_:.6f}")
print("\nCoefficients:")
for feat, coef in zip(features, reg.coef_):
    print(f"  {feat}: {coef:.6f}")

print(f"\nR-squared: {reg.score(X_normalized, y):.6f}")

# Regression equation
print("\nRegression Equation:")
eq_parts = [f"{reg.intercept_:.4f}"]
for feat, coef in zip(features, reg.coef_):
    eq_parts.append(f"({coef:.4f} * {feat})")
print(f"  Risk Factor = {' + '.join(eq_parts)}")

# Predictions
y_pred = reg.predict(X_normalized)
print("\nPredicted vs Actual Risk Factors:")
print(f"  {'Actual':>10} {'Predicted':>10} {'Difference':>10}")
for actual, pred in zip(y, y_pred):
    print(f"  {actual:>10.2f} {pred:>10.2f} {actual - pred:>10.4f}")


# =============================================================================
# Q2: SVM Classification
# =============================================================================
print("\n\n" + "=" * 70)
print("Q2: SVM Classification - Predict Risk Factor Relevance")
print("=" * 70)

# Prepare the dataset
# "Relevant" = risk factor above median threshold
median_risk = df_full_risk['Risk factor (L x I)'].median()
print(f"\nMedian Risk Factor: {median_risk:.2f}")
print("Risk Factor >= Median -> Relevant (1)")
print("Risk Factor <  Median -> Not Relevant (0)")

df_svm = df_full_risk.copy()

# Create binary target: 1 = relevant (high risk), 0 = not relevant (low risk)
df_svm['Relevant'] = (df_svm['Risk factor (L x I)'] >= median_risk).astype(int)

# Encode categorical features
df_svm['Event ID'] = df_svm['Event'].map(event_id_map)
df_svm['Sub-Event ID'] = range(1, len(df_svm) + 1)
df_svm['Risk Control Num'] = df_svm['Risk Control Present'].map({'Yes': 1, 'No': 0})

# Features for SVM: Sub-Event, Risk Control Present, Likelihood, Impact
svm_features = ['Sub-Event ID', 'Risk Control Num', 'Likelihood (L)', 'Impact (I)']
X_svm = df_svm[svm_features].values.astype(float)
y_svm = df_svm['Relevant'].values

# Normalize features
scaler_svm = MinMaxScaler()
X_svm_scaled = scaler_svm.fit_transform(X_svm)

# Split into train and test (70/30 split)
np.random.seed(42)
n = len(X_svm_scaled)
indices = np.random.permutation(n)
train_size = int(0.7 * n)
train_idx = indices[:train_size]
test_idx = indices[train_size:]

X_train = X_svm_scaled[train_idx]
X_test = X_svm_scaled[test_idx]
y_train = y_svm[train_idx]
y_test = y_svm[test_idx]

print(f"\nTraining set size: {len(X_train)}")
print(f"Test set size: {len(X_test)}")

# Train data
print("\nTraining Data (TechTrain):")
df_train = pd.DataFrame(X_train, columns=svm_features)
df_train['Relevant'] = y_train
print(df_train.to_string(index=False))

# Test data
print("\nTest Data (TechTest):")
df_test = pd.DataFrame(X_test, columns=svm_features)
df_test['Relevant'] = y_test
print(df_test.to_string(index=False))

# Train SVM
svm = SVC(kernel='rbf', C=1.0, gamma='scale', random_state=42)
svm.fit(X_train, y_train)

# Predict
y_pred_svm = svm.predict(X_test)

print("\n--- SVM Classification Results ---")
print(f"\nAccuracy: {accuracy_score(y_test, y_pred_svm):.4f}")

print("\nConfusion Matrix:")
cm = confusion_matrix(y_test, y_pred_svm)
print(f"  {'':>15} {'Pred: Not Rel':>15} {'Pred: Relevant':>15}")
print(f"  {'Act: Not Rel':>15} {cm[0][0] if cm.shape[0] > 0 else 0:>15} {cm[0][1] if cm.shape[0] > 0 and cm.shape[1] > 1 else 0:>15}")
if cm.shape[0] > 1:
    print(f"  {'Act: Relevant':>15} {cm[1][0]:>15} {cm[1][1] if cm.shape[1] > 1 else 0:>15}")

print("\nClassification Report:")
print(classification_report(y_test, y_pred_svm,
                            target_names=['Not Relevant', 'Relevant'],
                            zero_division=0))

# Predictions on full dataset
print("\nPredictions on Full Dataset:")
y_full_pred = svm.predict(X_svm_scaled)
df_predict = df_svm[['Event', 'Sub-Event', 'Risk factor (L x I)', 'Relevant']].copy()
df_predict['SVM Prediction'] = y_full_pred
df_predict['Correct'] = (df_predict['Relevant'] == df_predict['SVM Prediction'])
print(df_predict.to_string(index=False))
print(f"\nOverall Accuracy (full dataset): {accuracy_score(y_svm, y_full_pred):.4f}")

# =============================================================================
# Write results back to Excel
# =============================================================================
print("\n\n" + "=" * 70)
print("Writing results to Excel file: Week7-Risk_Assessment_Results.xlsx")
print("=" * 70)

with pd.ExcelWriter('Week7-Risk_Assessment_Results.xlsx', engine='openpyxl') as writer:
    # Risk Assessment table
    df_full_risk.to_excel(writer, sheet_name='Risk Assessment', index=False)

    # Normalized + Regression data
    df_reg_out = df_reg[['Event ID', 'Sub-Event ID', "Sub-Events' Counts", 'Probability',
                         'Risk Control Present Num', 'Likelihood (L)', 'Impact (I)',
                         'Risk factor (L x I)']].copy()
    df_reg_out.to_excel(writer, sheet_name='Normalized', index=False)

    # Regression coefficients
    reg_results = pd.DataFrame({
        'Feature': ['Intercept'] + features,
        'Coefficient': [reg.intercept_] + list(reg.coef_)
    })
    reg_results.to_excel(writer, sheet_name='Linear Regression', index=False)

    # Training data
    df_train_out = df_svm.iloc[train_idx][['Event', 'Sub-Event', 'Risk Control Num',
                                            'Likelihood (L)', 'Impact (I)', 'Relevant']].copy()
    df_train_out.to_excel(writer, sheet_name='TechTrain', index=False)

    # Test data
    df_test_out = df_svm.iloc[test_idx][['Event', 'Sub-Event', 'Risk Control Num',
                                          'Likelihood (L)', 'Impact (I)', 'Relevant']].copy()
    df_test_out['SVM Prediction'] = y_pred_svm
    df_test_out.to_excel(writer, sheet_name='TechTest', index=False)

    # Full predictions
    df_predict.to_excel(writer, sheet_name='Predict', index=False)

print("\nDone! Results saved to 'Week7-Risk_Assessment_Results.xlsx'")
